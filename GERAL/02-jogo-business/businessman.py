# Remove dados
#worksheet.delete_rows(2)
# Altera dados
#worksheet['A1'] = 'Maria do Carmos'

# Adiciona dados
#worksheet.append(['Novissmo', 'dado do novissimo'])

#importações
from unidecode import unidecode
from time import sleep
import openpyxl
from random import randint


#valor inicial do laço.
m = '0'
mbanco = True

#abrindo planilhas e definindo status do jogador
workbook = openpyxl.load_workbook('E:/Workspace/_CODE\PYTHON\GERAL/02-jogo-business/jogo1.xlsx')
jogador1 = workbook['jogador1']
mercado = workbook['mercado']
banco = workbook['banco']
negocios = workbook['negocios_jog']
locais = workbook['locais']
prod = workbook['produtos']
bolsa = workbook['bolsa']
caixa = jogador1.cell(row=1, column=2).value
score = jogador1.cell(row=2, column=2).value
lucro = jogador1.cell(row=3, column=2).value


#menu principal
while m != 'F':
    print(f'''STATUS:
    💰 CAIXA: {caixa} | 💎 {score} |  🚀 LUCRO: {lucro} | 💸 DÍVIDA: ''')
    print('''JOGO INICIADO!
    💻 [G] Gerenciador de negócios
    🛒 [M] Mercado
    🌏 [BV] Bolsa de valores
    🏦 [BAN] Banco
    ⏩ [AVA] Avançar
    ❌ [F] Sair do jogo''')
    m = str(input('Selecionar: ')).upper().strip()
    if m == 'G':
        menu = '0'                  
        while menu != 'V':
            print('=-='*10)
            print('Entrou no 💻 Gerenciador de negócios')
            print('=-='*10)
            print('\33[1;33m**MENU**\033[m')
            print('[L] Local\n[P] Pessoas\n[F] Ferramentas')
            print('\n**MINHAS EMPRESAS**')
            for row in negocios.iter_rows(min_row=2, values_only=True):
                    col1, col2, col3, col4, col5, col6, col7 = row
                    print(f'[{str(col1)}] | 🏢 {col2} | 💲 \033[31mR$ {col3}\033[m] | 💰 \033[32mR$ {col4}\033[m |  💎 {col5} | 🧔 {col6} | 📍 {col7}')                     
            menu = str(input('\033[34mSelecione [V] para sair:\033[m' )).upper().strip()
            if menu == 'L':                                
                print('*'*10)
                print('LOCAL')
                print('*'*10)
                print('\33[1;36m[C] Centro\n[L] Litoral\n[R] Rural\n[AI] Área industrial\n[AR] Área remota\033[m')                                
                menu = str(input('\033[34mSelecione [V] para sair:\033[m' )).upper().strip()
                if menu == 'C':
                    print('\033[1;33mBuscando novos locais...\033[m')
                    sleep(1)
                    print('\033[1;32mEncontrado!\033[m')
                    for row in locais.iter_rows(min_row=2, values_only=True):
                        col1, col2, col3, col4, col5 = row
                        print(f'[{str(col1)}] | 🏢 {col2} | 🧔 {col3} | 💲\033[31m{col4}\033[m | 📍{col5}')
                    chave = int(input('Digite o nº da empresa ou [0] para sair: '))  
                    encontrado = False                
            elif menu == 'P':
                print('*'*10)
                print('PESSOAS')
                print('*'*10)
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'F':
                print('*'*10)
                print('FERRAMENTAS')
                print('*'*10)
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'V':
                print('\033[31mSaindo do Gerenciadorg...\033[m')
                sleep(1)   
    elif m == 'M':
        print('-=-' * 10)
        print('🛒 MERCADO') 
        print('-=-' * 10)
        for row in mercado.iter_rows(min_row=2, values_only=True):
                col1, col2, col3, col4, col5, col6, col7 = row
                print(f'[{str(col1)}] | 🏢 {col2} | 💲 \033[31mR$ {col3}\033[m] | 💰 \033[32mR$ {col4}\033[m |  💎 {col5} | 🧔 {col6} | 📍 {col7}') 
        chave = int(input('Digite o nº da empresa ou [0] para sair: '))  
        encontrado = False    
        # Percorre as linhas da planilha em busca do nome do cliente
        for row in mercado.iter_rows(min_row=2, values_only=True):
            col1, col2, col3, col4, col5, col6, col7 = row            
            if chave == col1:                
                print('\033[1;33mTransação em andamento...\033[m')
                caixa = caixa - col3
                score = score + col5
                jogador1['B1'] = caixa
                jogador1['B2'] = score
                dados = [col1, col2, col3, col4, col5, col6, col7]
                negocios.append(dados)
                workbook.save('E:/Workspace/_CODE\PYTHON\GERAL/02-jogo-business/jogo1.xlsx')
                sleep(2)
                print('-=' * 10)
                print(f'Parabéns, você comprou \033[36m{col2}\033[m')
                sleep(1)
                encontrado = True
        if not encontrado:             
            print('-=' * 10)
            print('\033[31mEmpresa não identificada!\033[m')
            print('\033[32mSaindo do mercado...\033[m')
            sleep(1)
            print('-=' * 10)
    #GESTÃO DA BOLSA 
    elif m == 'BV':
        menu = '0' 
        print('=-='*10)
        print('🌏 BOLSA DE VALORES')
        print('=-='*10)
        while menu != 'V':
            print('[C] Comprar açõesl\n[S] vender ações')
            print('\n**EMPRESAS**')
            menu = str(input('\033[34mSelecione [V] para sair:\033[m' )).upper().strip()
            if menu == 'C':
                print('Compra de ações')
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'S':
                print('Venda de ações')
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'V':
                print('\033[32mSaindo da BV...\033[m')
                sleep(1)
           
    #GESTÃO DO BANCO
    elif m == 'BAN':
        menu = '0'                  
        while menu != 'V':
            print('=-='*10)
            print('Entrou no 🏦 BANCO')
            print('=-='*10)
            print('[E] Empréstimos\n[P] Poupança\n[D] Pagar dívida')                      
            menu = str(input('\033[34mSelecione [V] para sair:\033[m' )).upper().strip()
            if menu == 'E':                                
                print('*'*10)
                print('EMPRÉSTIMOS')
                print('*'*10)
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'P':
                print('*'*10)
                print('POUPANÇA')
                print('*'*10)
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'D':
                print('*'*10)
                print('PAGAMENTO DE DÍVIDAS')
                print('*'*10)
                menu = str(input('Selecione [0] para sair: '))
            elif menu == 'V':
                print('\033[31mSaindo do banco...\033[m')
                sleep(1)   
    elif m == 'AVA':
        print('Indisponível')
       
    elif m == 'F':
        print('Você saiu do jogo.')
        workbook.save('E:/Workspace/_CODE\PYTHON\GERAL/02-jogo-business/jogo1.xlsx')
print('FIM')

#PROBLEMAS
#como calcular e gerar renda dos negócios