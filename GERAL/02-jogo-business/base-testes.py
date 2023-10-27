from unidecode import unidecode
from time import sleep
import openpyxl
from random import randint


#valor inicial do laço.
m = '0'
mbanco = True

#abrindo planilhas e definindo status do jogador
workbook = openpyxl.load_workbook('E:/Workspace/_CODE\PYTHON\GERAL/02-jogo-business/jogo1.xlsx')
jogador1 = workbook['jogador1']
mercado = workbook['mercado']
banco = workbook['banco']
negocios = workbook['negocios_jog']
locais = workbook['locais']
prod = workbook['produtos']
caixa = jogador1.cell(row=1, column=2).value
score = jogador1.cell(row=2, column=2).value
lucro = jogador1.cell(row=3, column=2).value


print('AVANÇANDO...')
number_to_find = 1
column_to_find = 1 # por exemplo, coluna B
for row in prod.iter_rows(min_row=2, values_only=True):
        col1, col2, col3, col4, col5  = row
        print(f'[{col1}] | Vendas: {col2} | Produto: {col3} | Valor:{col4} | Tipo: {col5}')

        for row in prod.iter_rows():
            # Verifica se a célula contém o número inteiro desejado
            if row[col1 - 1] == number_to_find:
                # Gera um número aleatório e adiciona na célula ao lado        
                row[col1] = randint(1,5)

# Salva o arquivo com os novos valores
workbook.save('E:/Workspace/_CODE\PYTHON\GERAL/02-jogo-business/jogo1.xlsx')
print('finalizado!')



'''
print('AVANÇANDO...')
sleep(2)
for row in prod.iter_rows(min_row=2, values_only=True):
        col1, col2, col3, col4  = row
        print(f'[{col1}] | Tipo: {col2} | Produto: {col3} | Valor:{col4}') 
if col1 == 1:
    vendas = randint(500, 3500)
    lucro = vendas * col4
    caixa = caixa + lucro
    jogador1['B1'] = caixa
    print(f'{col2} vendeu {vendas} do produto {col3}')
elif col1 == 2:
    vendas = randint(0,3)
    lucro = vendas * col4
    caixa = caixa + lucro
    jogador1['B1'] = caixa
    print(f'{col2} vendeu {vendas} do produto {col3}')                    
workbook.save('E:/Workspace/_CODE\PYTHON\GERAL/02-jogo-business/jogo1.xlsx')
print('\033[32mFINALIZADO!\033[m')'''