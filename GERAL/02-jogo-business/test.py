import openpyxl
workbook = openpyxl.load_workbook('jogo1.xlsx')
geral = workbook['Planilha1']
jogador1 = workbook['jogador1']
mercado = workbook['mercado']
caixa = jogador1.cell(row=1, column=2).value
score = jogador1.cell(row=2, column=2).value
print(caixa)