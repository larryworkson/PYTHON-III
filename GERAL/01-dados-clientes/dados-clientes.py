# Remove dados
#worksheet.delete_rows(2)
# Altera dados
#worksheet['A1'] = 'Maria do Carmos'

# Adiciona dados
#worksheet.append(['Novissmo', 'dado do novissimo'])
# Salva as alterações na planilha
#workbook.save('clientes.xlsx')

from unidecode import unidecode
from time import sleep
import openpyxl

# Abre a planilha
workbook = openpyxl.load_workbook('E:/Workspace/_CODE\PYTHON\GERAL/01-dados-clientes/Clientes.xlsx')
# Seleciona a planilha a ser trabalhada
worksheet = workbook['Planilha1']


menu = -1
menu2 = -1
while menu != 0:
    print('\033[31m*\033[m' * 50)
    print('\033[1;34mO que deseja fazer?\033[m \n[1] Cadastrar cliente \n[2] Consultar clientes \n[0] Sair')
    menu = int(input('Insira sua opção: '))
    if menu == 1:
        print('\033[1;34mInsira os dados do cliente:\033[m')
        nm = str(input('Nome: ')).strip().upper()
        id = int(input('Idade: '))
        cd = str(input('Cidade: ')).strip().upper()
        worksheet.append([nm, id, cd])
        workbook.save('E:/Workspace/_CODE\PYTHON\GERAL/01-dados-clientes/Clientes.xlsx')
        print('\033[1;32mDados cadastrados com sucesso!\033[m')
        sleep(2)
    elif menu == 2:
        print('*' * 10)
        print('\033[35m[1] Pesquisar pelo nome\033[m \n\033[36m[2] Pesquisar pela cidade\n\033[m\033[34m[3] Consultar todos os cadastros\033[m\n[4] Cliente mais velho e mais novo')
        menu2 = int(input('Insira sua opção: '))
        if menu2 == 1:
            nome_procurado = str(input('Qual nome do cliente? ')).upper().strip()  
            encontrado = False    
            # Percorre as linhas da planilha em busca do nome do cliente
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                nome, idade, cidade = row
                primeiro_nome = nome.split()[0] #pega apenas o primeiro nome
                if unidecode(nome).find(unidecode(nome_procurado)) != -1:
                    # O nome do cliente foi encontrado, imprime os dados do cadastro
                    print('\033[1;33mBuscando...\033[m')
                    sleep(2)
                    print('-=' * 10)
                    print('\033[1;32mCliente encontrado!\033[m')
                    sleep(1)
                    print(f'Nome: {nome} | Idade: {idade} | Cidade: {cidade}')
                    encontrado = True
                    print('\033[1;m[1] Editar cadastro!\n[2] Excluir cadastro\033[m')

            if not encontrado:
                print('\033[1;33mBuscando...\033[m')
                sleep(2)                    
                print('-=' * 10)
                print('\033[1;31mCliente não encontrado!\033[m')
                print('-=' * 10)
        elif menu2 == 2:
            cidade_desejada = str(input('Nome da cidade: ')).upper().strip()
            clientes_encontrados = []
            print('\033[1;33mBuscando...\033[m')
            sleep(2)  
            for linha in worksheet.iter_rows(min_row=2, values_only=True):
                nome, idade, cidade = linha
                if unidecode(cidade) == unidecode(cidade_desejada):
                    clientes_encontrados.append({'nome': nome, 'idade': idade, 'cidade': cidade})
            if len(clientes_encontrados) == 0:
                print(f'Nenhum cadastro encontrado na cidade de {cidade_desejada}')
            else:
                for i, cliente in enumerate(clientes_encontrados):
                    print(f"Nome: {cliente['nome']} | Idade: {cliente['idade']} | Cidade: {cliente['cidade']}")
        elif menu2 == 3:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                nome, idade, cidade = row
                print(f'Nome: {nome} | Idade: {idade} | Cidade: {cidade}')                  
print('FIM')
#refirnar a pesquisa da cidade, para encontrar mesmo que o usuário digite parte do nome. Ex: Navegante (sem o S)
#colocar pesquisa cliente mais velho, mais novo


